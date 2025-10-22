import os
import json
import streamlit as st
from textwrap import dedent

from agno.agent import Agent
from agno.models.google import Gemini
from agno.tools.file import FileTools

from openpyxl import load_workbook
from openpyxl.comments import Comment
import pandas as pd

# ---------------- Utility Functions ----------------
def convert_to_csv(file_path:str):
   """
    Use this tool to convert the excel file to CSV.

    * file_path: Path to the Excel file to be converted
    """
   # Load the file  
   df = pd.read_excel(file_path).head(10)

   # Convert to CSV
   st.write("Converting to CSV... :leftwards_arrow_with_hook:")
   return df.to_csv('temp.csv', index=False)

def add_comments_to_header(file_path: str, data_dict: str = "data_dict.json"):
    """
    Add data dictionary as comments to header row of each sheet in Excel.
    """
    if not os.path.exists(data_dict):
        st.warning("Data dictionary not found, skipping comments.")
        return

    try:
        with open(data_dict, "r",errors="ignore") as f:
            data_dict_json = json.load(f)

        wb = load_workbook(file_path)

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for n, col in enumerate(ws.iter_cols(min_row=1, max_row=1)):
                for header_cell in col:
                    header_cell.comment = Comment(dedent(f"""
                        ColName: {data_dict_json[str(n)]['ColName']}, 
                        DataType: {data_dict_json[str(n)]['DataType']},
                        Description: {data_dict_json[str(n)]['Description']}
                    """), 'AI Agent')

        st.write("Saving File with comments... :floppy_disk:")
        wb.save('output.xlsx')
        with open('output.xlsx', 'rb') as f:
            st.download_button(
                label="Download output.xlsx",
                data=f,
                file_name='output.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    except Exception as e:
        st.error(f"Error adding comments: {e}")

# ---------------- Create Agent ----------------
def create_agent(api_key):
    """
    Create a Gemini agent to handle multi-sheet Excel modifications based on user instructions.
    """
    agent = Agent(
        model=Gemini(id="gemini-2.0-flash", api_key=api_key),
        description = dedent("""\
            You are a smart Excel agent that works with complex Excel files containing one or more sheets. 
            Your main task is to update the Excel sheets based on user prompts while strictly preserving the original format.

            Your responsibilities:

            1. Read the Excel workbook and understand the structure of each sheet.
            2. Accept user instructions in natural language describing changes. Example instructions:
            - "Update the date and items in this invoice with new data"
            - "Apply 20 percent discount to the final cost"
            - "Change employee salary and recalculate deductions in the payslip"
            3. Automatically perform any calculations required by the prompt:
            - Totals, subtotals
            - Taxes (GST, VAT, etc.)
            - Discounts or other adjustments
            4. Ensure that all sheets and all cells in the workbook maintain the original formatting, layout, formulas, and style.
            5. For multi-sheet workbooks, apply changes consistently across sheets if needed.
            6. Output a modified Excel file reflecting all changes.
            7. Optionally, provide a data dictionary or summary of changes if requested.

            Constraints:
            - Do not modify the format beyond what the user instructed.
            - Preserve formulas, styling, and layout unless changes require recalculation.
            - Any missing or unclear information should be inferred logically where possible.
            """),
        tools=[ FileTools() ],
        retries=2,
        # show_tool_calls=True
        )
    return agent

# ---------------- Streamlit UI ----------------
st.set_page_config(layout="centered", page_title="SheetShift", page_icon=":page_with_curl:")
st.title("SheetShift: Multi-Sheet Excel Automation Agent")
st.subheader("Modify Excel sheets dynamically based on user instructions")

with st.sidebar:
    api_key = st.text_input("Enter Gemini API key", type="password")
    input_file = st.file_uploader("Upload Excel file (.xlsx)", type='xlsx')
    user_prompt = st.text_area(
        "Enter your instructions for Excel changes:",
        placeholder="E.g., Update invoice items, apply discount, recalc totals, modify payslips..."
    )
    agent_run = st.button("Run Agent")
    if st.button("Reset Session"):
        st.session_state.clear()
        st.rerun()

# ---------------- Run Agent ----------------
if agent_run and input_file and user_prompt:
    file_path = "uploaded.xlsx"
    with open(file_path, "wb") as f:
        f.write(input_file.getbuffer())

    
    # csv_path = convert_to_csv(file_path)
    print("File uploaded and converted to CSV.")
    st.progress(15, text="Processing CSV...")

    
    if not os.path.exists(file_path):
        st.error(f"CSV file not found at {file_path}. Conversion might have failed.")
    else:
        agent = create_agent(api_key)
        st.write("Running Agent... :runner:")
        st.progress(50, text="AI Agent is running...")

        
        agent.print_response(dedent(f"""
            1. Read all sheets from '{file_path}' using openpyxl (do not attempt to read as text).
            2. Apply only the instructions provided by the user:
            {user_prompt}
            3. Perform any calculations (totals, taxes, discounts) only if explicitly mentioned.
            4. Preserve all formatting, formulas, layout, and style.
            5. Generate updated response in json format.
        """), markdown=True)
        st.progress(80, text="Finalizing changes...")           
    
    if os.path.exists('modified_output.xlsx'):
        with open('modified_output.xlsx', 'rb') as f:
            st.download_button(
                label="Download modified Excel",
                data=f,
                file_name="modified_output.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    
    add_comments_to_header(file_path, "data_dict.json")

    
    for tmp_file in ["temp.csv", "data_dict.json", file_path]:
        if os.path.exists(tmp_file):
            os.remove(tmp_file)

    st.progress(100, text="Done!")


